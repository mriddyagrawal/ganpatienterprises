"""
Jio auto-refill report importer.

Reads a CSV / TSV / XLSX export from the Jio distributor portal
(jioconnect → DSM Reports → Order Details) and turns each row into a
`Sale` in our books, attributed to the correct salesman by FOS ID and
to the correct retailer by Partner PRM ID.

Two-step admin flow:

1. **Upload + Preview** — parse the file, plan the changes, render a
   summary ("I will create N sales, M new retailers, K new salesmen,
   skip P duplicates"). Nothing touches the database yet.
2. **Confirm** — apply the plan inside a transaction. Re-uploading the
   same file is safe; the `jio_order_id` unique constraint and the
   importer's own duplicate check make it idempotent.

Business rules baked in:
- Ganpati gives every retailer a **3% incentive** — the Jio "Order
  Amount" is the face value (credits delivered); the retailer owes
  `face_value / 1.03`. See `PLAN.md §1` and `futureplans.md #9`.
- Only `Order Type = AUTO` rows with `Order Status = Completed` are
  imported. Non-auto / pending / failed rows are skipped with a
  recorded reason in the preview.
- Unknown retailers (new `Partner PRM ID`) are auto-created and
  assigned to the FOS on the row that introduced them. Admin can
  reassign anytime via Django Admin.
- Unknown salesmen (new `FOS ID`) are auto-created with `is_active=False`
  and no usable password — admin must explicitly enable them and set
  a password before they can log in. The user is created with a
  recognisable username `fos-<id>` and the name from the report.
"""

from __future__ import annotations

import base64
import csv
import io
from dataclasses import asdict, dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone

from .audit import log_change
from .models import AuditLog, Payment, Retailer, Sale, Visit


User = get_user_model()


# 3% incentive rate. See futureplans.md #9 for the case where this becomes
# configurable; for now it's a constant.
RETAILER_DISCOUNT = Decimal("1.03")

# Column-name aliases (Jio's headers occasionally have inconsistent
# whitespace — `Partner  PRM ID` has two spaces between Partner and PRM).
# We normalize headers (lowercase, collapse whitespace, replace " " with "_")
# before lookup, so any of these end up at the same key.
NORMALIZED_HEADERS = {
    "parent_prm_id": "parent_prm_id",
    "parent_name": "parent_name",
    "order_id": "order_id",
    "order_date": "order_date",
    "order_time": "order_time",
    "order_type": "order_type",
    "rpos_ref_no": "rpos_ref_no",
    "partner_prm_id": "partner_prm_id",
    "partner_name": "partner_name",
    "order_amount": "order_amount",
    "transfer_amount": "transfer_amount",
    "transfer_date": "transfer_date",
    "transfer_time": "transfer_time",
    "order_status": "order_status",
    "fos_id": "fos_id",
    "fos_name": "fos_name",
    "estel_ref": "estel_ref",
}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class JioImportRow:
    """One validated row, ready to become a Sale."""

    order_id: str
    partner_id: str
    partner_name: str
    fos_id: str
    fos_name: str
    face_value: Decimal
    amount: Decimal
    occurred_at: datetime

    def to_session_dict(self) -> dict[str, Any]:
        """JSON-safe shape for stashing in the session between
        upload-preview and confirm."""
        d = asdict(self)
        d["face_value"] = str(self.face_value)
        d["amount"] = str(self.amount)
        d["occurred_at"] = self.occurred_at.isoformat()
        return d

    @classmethod
    def from_session_dict(cls, d: dict[str, Any]) -> "JioImportRow":
        return cls(
            order_id=d["order_id"],
            partner_id=d["partner_id"],
            partner_name=d["partner_name"],
            fos_id=d["fos_id"],
            fos_name=d["fos_name"],
            face_value=Decimal(d["face_value"]),
            amount=Decimal(d["amount"]),
            occurred_at=datetime.fromisoformat(d["occurred_at"]),
        )


@dataclass
class ImportPlan:
    """Result of parsing + planning, before any DB writes."""

    rows: list[JioImportRow]
    sales_to_create: int = 0
    skipped_duplicates: int = 0
    new_retailers: dict[str, str] = field(default_factory=dict)  # partner_id -> name
    new_salesmen: dict[str, str] = field(default_factory=dict)  # fos_id -> name
    total_face_value: Decimal = Decimal("0")
    total_amount: Decimal = Decimal("0")
    parse_errors: list[str] = field(default_factory=list)


@dataclass
class ImportResult:
    """Result of applying a plan."""

    created_sales: int = 0
    created_retailers: int = 0
    created_salesmen: int = 0
    skipped_duplicates: int = 0
    total_amount: Decimal = Decimal("0")
    total_face_value: Decimal = Decimal("0")


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def _normalize_header(name: str) -> str:
    """`Partner  PRM ID` → `partner_prm_id`. Collapses runs of whitespace
    and lowercases so two-space typos in Jio's headers don't break us."""
    return "_".join(name.strip().lower().split())


def detect_format(content: bytes) -> str:
    """`'xlsx'` for true XLSX (PK zip header); `'text'` for anything else
    (CSV, TSV, possibly other delimited text)."""
    return "xlsx" if content[:4] == b"PK\x03\x04" else "text"


def parse_file_content(content: bytes) -> list[dict[str, str]]:
    """Parse raw file bytes (xlsx, csv, or tsv) into a list of dicts keyed
    by normalized header. Strips whitespace from every cell. Skips blank
    leading rows. Returns `[]` if the file has no data."""
    if detect_format(content) == "xlsx":
        return _parse_xlsx(content)
    return _parse_text(content)


def _parse_text(content: bytes) -> list[dict[str, str]]:
    """Auto-detect TSV vs CSV by looking at the first non-blank line."""
    text = content.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()
    while lines and not lines[0].strip():
        lines.pop(0)
    if not lines:
        return []
    delimiter = "\t" if "\t" in lines[0] else ","

    reader = csv.DictReader(io.StringIO("\n".join(lines)), delimiter=delimiter)
    return [
        {
            _normalize_header(k): (v.strip() if isinstance(v, str) else v)
            for k, v in row.items()
            if k is not None
        }
        for row in reader
    ]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_raw = [
        [("" if cell.value is None else str(cell.value)) for cell in row]
        for row in ws.iter_rows()
    ]
    while rows_raw and not any(c.strip() for c in rows_raw[0]):
        rows_raw.pop(0)
    if not rows_raw:
        return []
    header = [_normalize_header(c) for c in rows_raw[0]]
    return [
        dict(zip(header, [v.strip() for v in row], strict=False))
        for row in rows_raw[1:]
        if any(c.strip() for c in row)
    ]


# ---------------------------------------------------------------------------
# Row validation
# ---------------------------------------------------------------------------


class RowSkipped(Exception):
    """Raised for rows that are not eligible for import (non-AUTO, failed,
    missing required field, etc.)."""


def _parse_row(row: dict[str, str]) -> JioImportRow:
    """Convert a normalized row dict into a `JioImportRow`. Raises
    `RowSkipped` for rows that aren't AUTO/Completed or that have malformed
    fields."""
    status = row.get("order_status", "").strip()
    if status.lower() != "completed":
        raise RowSkipped(
            f"Order {row.get('order_id', '?')}: status={status!r}, not Completed"
        )

    order_type = row.get("order_type", "").strip()
    if order_type.upper() != "AUTO":
        raise RowSkipped(
            f"Order {row.get('order_id', '?')}: type={order_type!r}, not AUTO"
        )

    order_id = row.get("order_id", "").strip()
    partner_id = row.get("partner_prm_id", "").strip()
    fos_id = row.get("fos_id", "").strip()
    if not (order_id and partner_id and fos_id):
        raise RowSkipped(
            f"Order {order_id or '?'}: missing one of order_id/partner_prm_id/fos_id"
        )

    face_value_raw = row.get("order_amount", "").strip()
    try:
        face_value = Decimal(face_value_raw)
    except Exception as e:
        raise RowSkipped(f"Order {order_id}: bad order_amount {face_value_raw!r} ({e})")

    if face_value <= 0:
        raise RowSkipped(f"Order {order_id}: non-positive order_amount {face_value}")

    amount = (face_value / RETAILER_DISCOUNT).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )

    date_str = row.get("order_date", "").strip()
    time_str = row.get("order_time", "").strip().zfill(6)
    try:
        dt = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H%M%S")
    except ValueError as e:
        raise RowSkipped(f"Order {order_id}: bad date/time {date_str!r} {time_str!r} ({e})")

    occurred_at = timezone.make_aware(dt, timezone.get_current_timezone())

    return JioImportRow(
        order_id=order_id,
        partner_id=partner_id,
        partner_name=row.get("partner_name", "").strip(),
        fos_id=fos_id,
        fos_name=row.get("fos_name", "").strip(),
        face_value=face_value,
        amount=amount,
        occurred_at=occurred_at,
    )


def validate_rows(raw_rows: list[dict[str, str]]) -> tuple[list[JioImportRow], list[str]]:
    """Convert raw row dicts → JioImportRows. Returns (rows, errors)
    where errors is the human-readable list of rows that were skipped
    (for display in the preview)."""
    rows, errors = [], []
    for raw in raw_rows:
        try:
            rows.append(_parse_row(raw))
        except RowSkipped as e:
            errors.append(str(e))
    return rows, errors


# ---------------------------------------------------------------------------
# Planning
# ---------------------------------------------------------------------------


def plan_import(rows: list[JioImportRow]) -> ImportPlan:
    """Compute what the apply step will do, without touching the DB."""
    if not rows:
        return ImportPlan(rows=[])

    order_ids = [r.order_id for r in rows]
    partner_ids = [r.partner_id for r in rows]
    fos_ids = [r.fos_id for r in rows]

    existing_order_ids = set(
        Sale.objects.filter(jio_order_id__in=order_ids).values_list("jio_order_id", flat=True)
    )
    existing_partner_ids = set(
        Retailer.objects.filter(jio_partner_id__in=partner_ids).values_list(
            "jio_partner_id", flat=True
        )
    )
    existing_fos_ids = set(
        User.objects.filter(jio_fos_id__in=fos_ids).values_list("jio_fos_id", flat=True)
    )

    plan = ImportPlan(rows=rows)
    for r in rows:
        if r.order_id in existing_order_ids:
            plan.skipped_duplicates += 1
            continue
        plan.sales_to_create += 1
        plan.total_face_value += r.face_value
        plan.total_amount += r.amount
        if r.partner_id not in existing_partner_ids and r.partner_id not in plan.new_retailers:
            plan.new_retailers[r.partner_id] = r.partner_name
        if r.fos_id not in existing_fos_ids and r.fos_id not in plan.new_salesmen:
            plan.new_salesmen[r.fos_id] = r.fos_name
    return plan


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------


@transaction.atomic
def apply_plan(plan: ImportPlan, actor) -> ImportResult:
    """Create unknown retailers + salesmen, then create Sales. Atomic —
    a failure anywhere rolls the whole import back."""
    # 1. Auto-create unknown salesmen.
    fos_user_map: dict[str, Any] = {}
    for fos_id, name in plan.new_salesmen.items():
        u = User.objects.create(
            username=f"fos-{fos_id}",
            full_name=name,
            role=User.Role.SALESMAN,
            jio_fos_id=fos_id,
            is_active=False,  # admin must enable + set a password
        )
        u.set_unusable_password()
        u.save(update_fields=["password"])
        fos_user_map[fos_id] = u
    for u in User.objects.filter(jio_fos_id__in=[r.fos_id for r in plan.rows]).exclude(
        jio_fos_id__in=plan.new_salesmen
    ):
        fos_user_map[u.jio_fos_id] = u

    # 2. Auto-create unknown retailers.
    partner_retailer_map: dict[str, Any] = {}
    for partner_id, name in plan.new_retailers.items():
        # Assigned salesman = the FOS on the first row that introduced this retailer.
        first_fos_id = next(r.fos_id for r in plan.rows if r.partner_id == partner_id)
        retailer = Retailer.objects.create(
            name=name,
            jio_partner_id=partner_id,
            assigned_salesman=fos_user_map[first_fos_id],
        )
        partner_retailer_map[partner_id] = retailer
    for r_obj in Retailer.objects.filter(
        jio_partner_id__in=[r.partner_id for r in plan.rows]
    ).exclude(jio_partner_id__in=plan.new_retailers):
        partner_retailer_map[r_obj.jio_partner_id] = r_obj

    # 3. Create the Sales (skipping any whose jio_order_id already exists).
    existing_order_ids = set(
        Sale.objects.filter(jio_order_id__in=[r.order_id for r in plan.rows]).values_list(
            "jio_order_id", flat=True
        )
    )

    result = ImportResult(
        created_retailers=len(plan.new_retailers),
        created_salesmen=len(plan.new_salesmen),
        skipped_duplicates=plan.skipped_duplicates,
    )
    for r in plan.rows:
        if r.order_id in existing_order_ids:
            continue
        sale = Sale(
            retailer=partner_retailer_map[r.partner_id],
            salesman=fos_user_map[r.fos_id],
            amount=r.amount,
            face_value=r.face_value,
            occurred_at=r.occurred_at,
            jio_order_id=r.order_id,
        )
        # AUTO refills aren't physical visits — skip the visit_attach side
        # effect so the Visit table only contains real salesman activity.
        sale.save(skip_visit_attach=True)
        log_change(actor=actor, instance=sale, action=AuditLog.Action.CREATE)
        result.created_sales += 1
        result.total_amount += r.amount
        result.total_face_value += r.face_value
    return result


# ---------------------------------------------------------------------------
# Session helpers — bytes round-trip for the preview→confirm dance
# ---------------------------------------------------------------------------


def stash_file_content(content: bytes) -> str:
    """Base64-encode bytes for stashing in the session between the
    upload-preview and the confirm step."""
    return base64.b64encode(content).decode("ascii")


def unstash_file_content(stashed: str) -> bytes:
    return base64.b64decode(stashed.encode("ascii"))
